import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import time

excel_file_path = r"Excel.xlsx"
chrome_driver_path = r"chromedriver.exe"

def initialize_web_driver():
    os.environ["PATH"] += os.pathsep + chrome_driver_path
    driver = webdriver.Chrome()
    driver.get("https://www.google.com")
    return driver

def get_suggestions_from_google(driver, keyword):
    search_box = driver.find_element(By.NAME, "q")
    search_box.clear()
    search_box.send_keys(keyword)

    wait = WebDriverWait(driver, 10)
    wait.until(EC.presence_of_element_located((By.XPATH, "//ul[@role='listbox']/li[@role='presentation']")))
    time.sleep(2)  # Adding a small delay to ensure suggestions load completely

    suggestions = driver.find_elements(By.XPATH, "//ul[@role='listbox']/li[@role='presentation']")
    suggestion_texts = [suggestion.text for suggestion in suggestions]

    if suggestion_texts:
        longest_suggestion = max(suggestion_texts, key=len)
        shortest_suggestion = min(suggestion_texts, key=len)
        return longest_suggestion, shortest_suggestion
    else:
        return None, None

def main():
    workbook = openpyxl.load_workbook(excel_file_path)

    # Get the current day
    current_day = datetime.now().strftime("%A")

    if current_day in workbook.sheetnames:
        worksheet = workbook[current_day]

        longest_suggestions = []
        shortest_suggestions = []
        web_driver = initialize_web_driver()

        # Iterate through the rows and get suggestions
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=2):
            keyword = row[0].value
            if keyword and not keyword.isspace():
                print("Getting suggestions for:", keyword)
                long_suggestion, short_suggestion = get_suggestions_from_google(web_driver, keyword)
                print("Longest Suggestion:", long_suggestion)
                print("Shortest Suggestion:", short_suggestion)
                longest_suggestions.append(long_suggestion)
                shortest_suggestions.append(short_suggestion)
                time.sleep(2)  # Adding a wait time between each keyword

        for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=4)):
            row[0].value = longest_suggestions[idx]
            row[1].value = shortest_suggestions[idx]

        # Save the updated Excel file
        workbook.save(excel_file_path)

        workbook.close()
        web_driver.quit()

if __name__ == "__main__":
    main()
